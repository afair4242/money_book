import streamlit as st
import openpyxl as op
import io
import json
import os
import pandas as pd

st.title("엑셀 정산 프로그램")
st.write('개인농협+회사농협+하나카드 엑셀준비 (다운로드 후 다른이름으로 저장 xlsx)')
st.write('홈페이지 제작 또는 기타 개인수익과 카드에 포함되지 않는 항목은 수동입력해야함')

tab1, tab2, tab3 = st.tabs(["호스팅수입(회사)", "배달수입(개인)", "신용카드 지출정산"])

with tab1:
    #st.subheader("회사수익(호스팅) 정산")
    uploaded_file = st.file_uploader("회사엑셀 파일을 업로드 하세요.", type=["xlsx"],key="file_uploader_1")
    st.write('---')
    if uploaded_file is not None:
        try:
            # 업로드된 파일을 메모리에서 읽기
            wb = op.load_workbook(io.BytesIO(uploaded_file.getvalue()), data_only=True)
            ws = wb.active  # 활성화된 시트 선택

            # 합산할 값 목록
            target_values = {66000, 88000, 385000}
            total_sum = 0
            included_items = []  # 포함된 항목 리스트

            # E열(금액)과 G열(설명) 탐색
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=7):  # E, G 컬럼 탐색
                amount = row[0].value  # E열: 금액
                description = row[2].value  # G열: 설명

                if isinstance(amount, (int, float)) and amount in target_values:
                    total_sum += amount
                    included_items.append({"항목": description, "금액": amount})

            # 결과 출력
            st.subheader(f"호스팅 합산: {total_sum} 원")

            # 포함된 항목 테이블 출력
            if included_items:
                st.caption("📌 포함된 항목 목록")
                df = pd.DataFrame(included_items)
                st.table(df)

        except Exception as e:
            st.error(f"파일을 읽는 중 오류가 발생했습니다: {e}")

with tab2:
    #st.subheader("배달수입(개인) 정산")
    # 엑셀 파일 업로드 UI
    uploaded_file = st.file_uploader("개인엑셀 파일을 업로드하세요", type=["xlsx"], key="file_uploader_2")
    st.write('---')
    if uploaded_file is not None:
        try:
            # 업로드된 파일을 메모리에서 읽기
            wb = op.load_workbook(io.BytesIO(uploaded_file.getvalue()), data_only=True)
            ws = wb.active  # 활성화된 시트 선택

            # 정산 금액 초기화
            coupang_total = 0
            baemin_total = 0

            # 정산 내역 저장 리스트
            coupang_details = []
            baemin_details = []

            # 데이터 탐색 (C열: 날짜, E열: 금액, F열: 내용)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=6):  # C, E, F 컬럼 탐색
                date = row[0].value  # C열: 날짜
                amount = row[2].value  # E열: 금액
                description = row[3].value  # F열: 내용

                # 값이 숫자인 경우만 처리
                if isinstance(amount, (int, float)):
                    if description == "쿠팡이츠정산":
                        coupang_total += amount
                        coupang_details.append({"날짜": date, "금액": amount, "내용": description})
                    elif description == "우아한청년들":
                        baemin_total += amount
                        baemin_details.append({"날짜": date, "금액": amount, "내용": description})

            # 전체 합계 계산
            overall_total = coupang_total + baemin_total

            # 결과 출력
            st.subheader(f"쿠팡정산 합계: {coupang_total} 원")
            if coupang_details:
                st.caption("📌 쿠팡이츠정산 내역")
                df_coupang = pd.DataFrame(coupang_details)
                st.table(df_coupang)

            st.subheader(f"배민정산 합계: {baemin_total} 원")
            if baemin_details:
                st.caption("📌 배민정산 내역")
                df_baemin = pd.DataFrame(baemin_details)
                st.table(df_baemin)

            # 전체 합계 출력
            st.markdown("---")
            st.subheader(f"💰 전체 합계: {overall_total} 원")

        except Exception as e:
            st.error(f"파일을 읽는 중 오류가 발생했습니다: {e}")

with tab3:

    #st.subheader("신용카드 지출정산")
    # 엑셀 파일 업로드 UI
    uploaded_file = st.file_uploader("신용카드 엑셀 파일을 업로드하세요", type=["xlsx"],key="file_uploader_3")
    st.write('---')

    # JSON 파일 경로
    JSON_FILE = "categories.json"

    # 기본 카테고리 목록
    categories = ["식비", "간식", "주유", "물품", "도메인", "호스팅", "구독", "기타",
                "통신비", "메리츠", "건강보험", "전기세"]

    # JSON 파일 로드 또는 생성
    if os.path.exists(JSON_FILE):
        with open(JSON_FILE, "r", encoding="utf-8") as f:
            category_keywords = json.load(f)
    else:
        category_keywords = {category: [] for category in categories}
        with open(JSON_FILE, "w", encoding="utf-8") as f:
            json.dump(category_keywords, f, ensure_ascii=False, indent=4)


    # 사용자 입력을 받아 각 카테고리에 해당하는 키워드 저장 (textarea 높이 2줄)
    updated_keywords = {}
    for category in categories:
        keywords = st.text_input(f"{category} 카테고리에 포함될 키워드 (쉼표로 구분)", 
                                ", ".join(category_keywords.get(category, [])))
        updated_keywords[category] = [kw.strip() for kw in keywords.split(",") if kw.strip()]

    # 변경된 키워드를 JSON 파일에 저장
    if st.button("키워드 저장"):
        with open(JSON_FILE, "w", encoding="utf-8") as f:
            json.dump(updated_keywords, f, ensure_ascii=False, indent=4)
        st.success("키워드가 저장되었습니다.")



    if uploaded_file is not None:
        try:
            # 엑셀 파일 읽기
            wb = op.load_workbook(io.BytesIO(uploaded_file.getvalue()), data_only=True)
            ws = wb.active  # 활성화된 시트 선택

            # 카테고리별 합계 저장용 딕셔너리 및 세부 항목 리스트 초기화
            category_totals = {category: 0 for category in categories}
            category_details = {category: [] for category in categories}
            unclassified_items = []  # 미분류 항목 저장 리스트
            unclassified_total = 0  # 미분류 항목 합계
            overall_total = 0  # 전체 합계

            # E열(설명)과 F열(금액) 탐색
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=6):  # E, F 컬럼 탐색
                description = row[0].value  # E열: 설명
                amount = row[1].value  # F열: 금액

                if isinstance(amount, (int, float)) and isinstance(description, str):
                    matched = False  # 항목이 분류되었는지 여부
                    overall_total += amount  # 전체 합계 누적

                    for category, keywords in updated_keywords.items():
                        if any(keyword in description for keyword in keywords):
                            category_totals[category] += amount
                            category_details[category].append({"항목": description, "금액": amount})
                            matched = True
                            break

                    if not matched:
                        unclassified_items.append({"항목": description, "금액": amount})
                        unclassified_total += amount

            # 결과 출력
            for category, total in category_totals.items():
                st.subheader(f"{category} 합계: {total} 원")
                if category_details[category]:
                    st.caption("📌 해당 카테고리에 포함된 항목:")
                    df = pd.DataFrame(category_details[category])
                    st.table(df)

            # 미분류 항목 출력
            if unclassified_items:
                st.subheader(f"❗ 미분류 항목 합계: {unclassified_total} 원")
                df_unclassified = pd.DataFrame(unclassified_items)
                st.table(df_unclassified)

            # 전체 합계 출력
            st.markdown("---")
            st.subheader(f"💰 전체 합계: {overall_total} 원")

        except Exception as e:
            st.error(f"파일을 읽는 중 오류가 발생했습니다: {e}")